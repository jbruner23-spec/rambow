#!/usr/bin/env python3
"""Seed / re-import Rambow inventory from MyFootballCards.xlsx into Supabase.

Sends the parsed rows to the `rmb_seed_inventory` RPC (SECURITY DEFINER), which
upserts players, rebuilds card_sets, and refreshes owned cards in one call. Uses
the publishable/anon key — no service-role secret required. This is both the
first-seed path and the ongoing re-import path (UC-01 / UC-02).

Env:  VITE_SUPABASE_URL, VITE_SUPABASE_ANON_KEY, RMB_SEED_TOKEN  (from ~/rambow/.env)

Run:  set -a; . ~/rambow/.env; set +a
      uv run --with requests --with openpyxl python scripts/import_inventory.py
"""
from __future__ import annotations
import os
import re
import sys
import openpyxl
import requests

XLSX = os.path.expanduser(
    "~/Library/Mobile Documents/com~apple~CloudDocs/Documents/MyFootballCards.xlsx")
SHEETS = ["Rams", "Faulk"]
FEATURED = {"Aaron Donald", "Marshall Faulk"}
JERSEY = {"Aaron Donald": "99", "Marshall Faulk": "28", "Cooper Kupp": "10",
          "Matthew Stafford": "9", "Kurt Warner": "13", "Torry Holt": "88",
          "Isaac Bruce": "80", "Eric Dickerson": "29", "Terrell Davis": "30"}
COL = dict(year=0, product=1, card_no=2, name=3, card_type=4,
           graded=5, serial=6, qty=7, price=8)

URL = (os.environ.get("VITE_SUPABASE_URL") or os.environ.get("SUPABASE_URL", "")).rstrip("/")
KEY = os.environ.get("VITE_SUPABASE_ANON_KEY") or os.environ.get("SUPABASE_ANON_KEY", "")
TOKEN = os.environ.get("RMB_SEED_TOKEN", "")
if not (URL and KEY):
    sys.exit("VITE_SUPABASE_URL / VITE_SUPABASE_ANON_KEY not set (source ~/rambow/.env)")
if not TOKEN:
    sys.exit("RMB_SEED_TOKEN not set (source ~/rambow/.env)")


def canon_product(p: str) -> str:
    return re.sub(r"^Panini\s+", "", str(p).strip())


def as_int(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    m = re.search(r"-?\d+", str(v))
    return int(m.group()) if m else None


def as_price(v):
    return round(float(v), 2) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def read_rows():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    rows, players = [], set()
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            continue
        for r in wb[sheet].iter_rows(values_only=True):
            if not r or all(c is None for c in r):
                continue
            name = r[COL["name"]]
            if not name or str(name).strip() in ("Name", ""):
                continue
            year, product = as_int(r[COL["year"]]), r[COL["product"]]
            if year is None or not product:
                continue
            player = str(name).strip()
            players.add(player)
            raw_cn = r[COL["card_no"]]
            if raw_cn is None:
                card_no = ""
            elif isinstance(raw_cn, float) and raw_cn.is_integer():
                card_no = str(int(raw_cn))          # 12.0 -> "12", not "12.0"
            else:
                card_no = str(raw_cn).strip()
            note = None
            if len(r) > 12 and isinstance(r[12], str) and r[12].strip().startswith("·"):
                note = re.sub(r"\s+", " ", r[12]).strip("· ").strip() or None
            rows.append(dict(
                player=player, year=year, product=canon_product(product), card_no=card_no,
                parallel=(str(r[COL["card_type"]]).strip() if r[COL["card_type"]] else None),
                graded=(str(r[COL["graded"]]).strip() if r[COL["graded"]] else None),
                serial=as_int(r[COL["serial"]]), qty=as_int(r[COL["qty"]]),
                price=as_price(r[COL["price"]]), notes=note))
    return rows, players


def main():
    rows, players = read_rows()
    payload = {
        "token": TOKEN,
        "players": [
            {"name": p, "featured": p in FEATURED, "jersey_no": JERSEY.get(p),
             "sort_order": (i if p in FEATURED else 10 + i)}
            for i, p in enumerate(sorted(players, key=lambda x: (x not in FEATURED, x)))],
        "rows": rows,
    }
    r = requests.post(f"{URL}/rest/v1/rpc/rmb_seed_inventory",
                      headers={"apikey": KEY, "Authorization": f"Bearer {KEY}",
                               "Content-Type": "application/json"},
                      json={"payload": payload}, timeout=90)
    if not r.ok:
        sys.exit(f"seed RPC failed {r.status_code}: {r.text[:500]}")
    print("seeded:", r.json(), f"(sent {len(rows)} rows)")


if __name__ == "__main__":
    main()
